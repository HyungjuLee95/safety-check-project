import base64
import io
import os
import re
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from copy import copy
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


def _repair_xlsx_style_ids(src_path: str) -> str:
    """Repair broken style index references in a template workbook."""
    tmpdir = tempfile.mkdtemp(prefix="xlsx_repair_")

    with zipfile.ZipFile(src_path, "r") as zf:
        zf.extractall(tmpdir)

    styles_path = os.path.join(tmpdir, "xl", "styles.xml")
    if not os.path.exists(styles_path):
        return src_path

    tree = ET.parse(styles_path)
    root = tree.getroot()
    ns_uri = root.tag.split("}")[0].strip("{")
    ns = {"a": ns_uri}

    cell_xfs = root.find("a:cellXfs", ns)
    xf_count = len(cell_xfs.findall("a:xf", ns)) if cell_xfs is not None else 0
    if xf_count <= 0:
        return src_path

    sheets_dir = os.path.join(tmpdir, "xl", "worksheets")
    if not os.path.isdir(sheets_dir):
        return src_path

    style_pattern = re.compile(r'\bs="(\d+)"')

    def _patch_sheet(sheet_xml_path: str) -> None:
        txt = open(sheet_xml_path, "r", encoding="utf-8").read()

        def _repl(match: re.Match[str]) -> str:
            style_id = int(match.group(1))
            return 's="0"' if style_id >= xf_count else match.group(0)

        patched = style_pattern.sub(_repl, txt)
        if patched != txt:
            with open(sheet_xml_path, "w", encoding="utf-8") as fp:
                fp.write(patched)

    for filename in os.listdir(sheets_dir):
        if filename.endswith(".xml"):
            _patch_sheet(os.path.join(sheets_dir, filename))

    repaired_path = os.path.join(tempfile.gettempdir(), f"repaired_{os.path.basename(src_path)}")
    if os.path.exists(repaired_path):
        os.remove(repaired_path)

    with zipfile.ZipFile(repaired_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for folder, _, files in os.walk(tmpdir):
            for file in files:
                full = os.path.join(folder, file)
                rel = os.path.relpath(full, tmpdir)
                zf.write(full, rel)

    return repaired_path


def _load_template_safe(template_path: str):
    try:
        return load_workbook(template_path)
    except IndexError:
        repaired = _repair_xlsx_style_ids(template_path)
        return load_workbook(repaired)


def _excel_col_width_to_pixels(width: Optional[float]) -> int:
    if width is None:
        width = 8.43
    return int(width * 7 + 5)


def _points_to_pixels(points: Optional[float]) -> int:
    if points is None:
        points = 15.0
    return int(points * 4 / 3)


def _estimate_range_pixels(ws, cell_range: str) -> Tuple[int, int]:
    start, end = cell_range.split(":")
    start_col = ws[start].column
    start_row = ws[start].row
    end_col = ws[end].column
    end_row = ws[end].row

    width_px = 0
    for col in range(start_col, end_col + 1):
        letter = get_column_letter(col)
        width_px += _excel_col_width_to_pixels(ws.column_dimensions[letter].width)

    height_px = 0
    for row in range(start_row, end_row + 1):
        height_px += _points_to_pixels(ws.row_dimensions[row].height)

    return width_px, height_px


def _find_merged_range_containing_cell(ws, coord: str) -> Optional[str]:
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            return merged_range.coord
    return None


def _decode_signature_image(signature_base64: Optional[str]) -> Optional[bytes]:
    if not signature_base64:
        return None

    raw = str(signature_base64).strip()
    if not raw:
        return None

    if raw.startswith("data:image/") and "," in raw:
        raw = raw.split(",", 1)[1]

    try:
        return base64.b64decode(raw, validate=False)
    except Exception:
        return None


def _insert_signature_image_pretty(ws, anchor_cell: str, image_bytes: Optional[bytes], padding_px: int = 2) -> None:
    if not image_bytes:
        return

    try:
        with PILImage.open(io.BytesIO(image_bytes)) as im:
            if im.mode in ("RGBA", "P"):
                im = im.convert("RGB")
            original_w, original_h = im.size
            render_buffer = io.BytesIO()
            im.save(render_buffer, format="PNG")
    except Exception:
        return

    original_w = max(1, int(original_w))
    original_h = max(1, int(original_h))

    render_buffer.seek(0)
    render_buffer.name = "signature.png"
    img = XLImage(render_buffer)

    merged = _find_merged_range_containing_cell(ws, anchor_cell)
    if merged:
        width_px, height_px = _estimate_range_pixels(ws, merged)
    else:
        col_letter = re.sub(r"\d+", "", anchor_cell)
        row_num = int(re.sub(r"\D+", "", anchor_cell))
        width_px = _excel_col_width_to_pixels(ws.column_dimensions[col_letter].width)
        height_px = _points_to_pixels(ws.row_dimensions[row_num].height)

    target_w = max(10, width_px - padding_px * 2)
    target_h = max(10, height_px - padding_px * 2)
    scale = min(target_w / original_w, target_h / original_h)

    img.width = max(10, int(original_w * scale))
    img.height = max(10, int(original_h * scale))

    ws.add_image(img, anchor_cell)


def _replace_placeholder(cell_value: Any, new_value: str) -> str:
    if not isinstance(cell_value, str):
        return new_value
    if "**(" in cell_value:
        return re.sub(r"\*\*\(.*?\)\*\*", new_value, cell_value)
    return new_value


def _make_unique_sheet_title(base_title: str, used_titles: set[str]) -> str:
    cleaned = re.sub(r"[\\/:?*\[\]]", "_", (base_title or "Sheet")).strip() or "Sheet"
    cleaned = cleaned[:31]

    if cleaned not in used_titles:
        used_titles.add(cleaned)
        return cleaned

    idx = 1
    while True:
        suffix = f"_{idx}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        if candidate not in used_titles:
            used_titles.add(candidate)
            return candidate
        idx += 1


def _copy_cell_style_and_format(src_cell, dst_cell) -> None:
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)


def _create_default_template_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    ws["A1"] = "작성일"
    ws["A2"] = "점검자"
    ws["A3"] = "병원"
    ws["A4"] = "작업종류"
    ws["C4"] = "장비"

    ws["B1"] = "**(작성일)**"
    ws["B2"] = "**(점검자)**"
    ws["B3"] = "**(병원)**"
    ws["B4"] = "**(작업종류)**"
    ws["D4"] = "**(장비)**"

    ws.merge_cells("E1:G1")
    ws.merge_cells("E2:G2")
    ws.merge_cells("E3:G3")
    ws.merge_cells("E4:G4")
    ws["E1"] = "이름: **(SUBADMIN)**"
    ws["E3"] = "이름: **(WORKER)**"

    ws["A6"] = "No"
    ws["B6"] = "질문"
    ws["C6"] = "결과"
    ws["D6"] = "코멘트"

    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 20 if col == "B" else 14

    ws.row_dimensions[2].height = 50
    ws.row_dimensions[4].height = 50

    return wb


def _write_record_to_sheet(
    ws,
    record: Dict[str, Any],
    name_row_height_pt: float = 18.0,
    sign_row_height_pt: float = 52.0,
) -> None:
    write_date = str(record.get("date") or "")
    inspector = str(record.get("userName") or record.get("name") or "")
    hospital = str(record.get("hospital") or "")
    work_type = str(record.get("workType") or "")
    equipment = str(record.get("equipmentName") or "")
    subadmin_name = str(record.get("subadminName") or "")
    worker_name = inspector

    ws["B1"].value = _replace_placeholder(ws["B1"].value, write_date)
    ws["B2"].value = _replace_placeholder(ws["B2"].value, inspector)
    ws["B3"].value = _replace_placeholder(ws["B3"].value, hospital)
    ws["B4"].value = work_type
    ws["C4"].value = _replace_placeholder(ws["C4"].value, equipment)

    ws.row_dimensions[1].height = name_row_height_pt
    ws.row_dimensions[2].height = sign_row_height_pt
    ws.row_dimensions[3].height = name_row_height_pt
    ws.row_dimensions[4].height = sign_row_height_pt

    ws["E1"].value = f"이름: {subadmin_name}" if subadmin_name else "이름:"
    ws["E3"].value = f"이름: {worker_name}" if worker_name else "이름:"

    ws["E2"].value = None
    ws["E4"].value = None

    _insert_signature_image_pretty(ws, "E2", _decode_signature_image(record.get("subadminSignatureBase64")), padding_px=2)
    _insert_signature_image_pretty(ws, "E4", _decode_signature_image(record.get("signatureBase64")), padding_px=2)

    # checklist rows
    results = record.get("results") or []
    start_row = 7
    for idx, result in enumerate(results, start=1):
        row = start_row + idx - 1
        ws[f"A{row}"] = idx
        ws[f"B{row}"] = result.get("question") or ""
        ws[f"C{row}"] = result.get("value") or ""
        ws[f"D{row}"] = result.get("comment") or ""


def _copy_template_sheet_content(src_ws, dst_ws) -> None:
    for row in src_ws.iter_rows():
        for src_cell in row:
            dst_cell = dst_ws.cell(row=src_cell.row, column=src_cell.column)
            dst_cell.value = src_cell.value
            _copy_cell_style_and_format(src_cell, dst_cell)

    for merged in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged))

    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width

    for row_idx, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_idx].height = dim.height


def build_inspections_excel_bytes(
    records: List[Dict[str, Any]],
    template_path: Optional[str] = None,
) -> bytes:
    if template_path and os.path.exists(template_path):
        template_wb = _load_template_safe(template_path)
    else:
        template_wb = _create_default_template_workbook()

    template_ws = template_wb["Sheet1"] if "Sheet1" in template_wb.sheetnames else template_wb.worksheets[0]

    wb = Workbook()
    wb.remove(wb.active)

    used_titles: set[str] = set()
    target_records = records or [{}]

    for record in target_records:
        base_name = f"{record.get('date') or 'NoDate'}_{record.get('name') or record.get('userName') or 'Inspection'}"
        title = _make_unique_sheet_title(base_name, used_titles)

        ws = wb.create_sheet(title=title)
        _copy_template_sheet_content(template_ws, ws)
        _write_record_to_sheet(ws, record)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_export_filename(start_date: str, end_date: str) -> str:
    safe_start = re.sub(r"[^0-9A-Za-z_-]", "_", start_date)
    safe_end = re.sub(r"[^0-9A-Za-z_-]", "_", end_date)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"safety_report_{safe_start}_to_{safe_end}_{timestamp}.xlsx"
